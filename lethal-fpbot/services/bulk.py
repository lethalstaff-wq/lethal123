"""Массовая загрузка лотов из CSV (опционально Excel через openpyxl).

Формат CSV:
    title,price,currency,description

Пример:
    Steam Random Key,50,RUB,Случайный ключ Steam
    GTA V Modded,500,RUB,Жирный аккаунт с миллиардами

Бот не пушит лоты в FunPay автоматически (это требует ручной модерации
у самого FunPay), а сохраняет их в auto_delivery как «черновики» —
дальше пользователь привязывает к настоящему лоту.
"""

from __future__ import annotations

import csv
import io
import logging
from collections.abc import Iterable

from database.models import add_auto_delivery

logger = logging.getLogger(__name__)


def parse_csv(text: str) -> list[dict]:
    """Парсит CSV-строку в список dict'ов с базовой валидацией."""
    out: list[dict] = []
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        title = (row.get("title") or "").strip()
        if not title:
            continue
        price = row.get("price") or row.get("Price") or "0"
        try:
            price_val = float(price)
        except (ValueError, TypeError):
            price_val = 0.0
        out.append(
            {
                "title": title,
                "price": price_val,
                "currency": (row.get("currency") or "RUB").strip().upper(),
                "description": (row.get("description") or "").strip(),
            }
        )
    return out


def parse_excel(content: bytes) -> list[dict]:
    """Парсит .xlsx через openpyxl. Если openpyxl не установлен — поднимет ImportError."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl required for excel parsing") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).lower().strip() if c else "" for c in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        record = dict(zip(headers, row, strict=False))
        title = (str(record.get("title") or "")).strip()
        if not title:
            continue
        try:
            price_val = float(record.get("price") or 0)
        except (ValueError, TypeError):
            price_val = 0.0
        out.append(
            {
                "title": title,
                "price": price_val,
                "currency": str(record.get("currency") or "RUB").upper(),
                "description": str(record.get("description") or ""),
            }
        )
    return out


async def import_lots_for_user(
    user_id: int, rows: Iterable[dict], items_per_lot: int = 10
) -> int:
    """Создаёт автовыдачу-черновики из распарсенных лотов.

    Каждый лот заводится с пустым items[] (или 1 заглушкой), пользователь
    дальше сам наполнит. Можно скастомизировать через items_per_lot.
    """
    count = 0
    for row in rows:
        await add_auto_delivery(
            user_id=user_id,
            fp_account_id=None,
            lot_name=row["title"],
            items=["DRAFT — заполните свои товары"] * max(1, items_per_lot),
            template=row.get("description") or None,
        )
        count += 1
    return count
